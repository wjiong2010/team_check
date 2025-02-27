import xml.dom.minidom
from xml.dom import DOMException
from kpi import KPIForOnePerson
from openpyxl import Workbook
from excel_format import ExcelFormat
from report import kpi_interview_form_generator

CR_ELEMENT_NODE = 1
xl_format = ExcelFormat()


def parse_work(m, node):
    for subNode in node.childNodes:
        if CR_ELEMENT_NODE == subNode.nodeType:
            if 'work_app' == subNode.nodeName:
                m.work_apps = subNode.getAttribute('item').split(',')
            elif 'work_module' == subNode.nodeName:
                m.work_modules = subNode.getAttribute('item').split(',')
            elif 'work_gl_app' == subNode.nodeName:
                m.work_gl_apps = subNode.getAttribute('item').split(',')
            elif 'work_pro_app' == subNode.nodeName:
                m.work_pro_apps = subNode.getAttribute('item').split(',')
            else:
                print("UNKNOWN node type: " + subNode.nodeName)


def kpi_head_text():
    report_head = '''任务量等级： 5:非常饱满， 4:饱满，3：适中，2：不足，1：严重不足'''
    report_main_seperator = '''================================================================================='''

    kpi_text = report_head + '\n'
    kpi_text += report_main_seperator + '\n\n'

    return kpi_text


class TeamMember:
    def cr_result_in_excel(self, ws, r):
        print(f"{str(r)} next line----------------------------")
        ori_r = r
        rc = f"A{r}"
        print(f"{rc}")
        ws[rc] = f"{self.name_cn}({self.name_en})".title()
        mg_rc = f"A{r}:B{r}"
        ws.merge_cells(mg_rc)
        xl_format.set_cell(ws[rc], font=xl_format.cell_title_font)
        rc = f"B{r}"
        xl_format.set_cell(ws[rc], font=xl_format.cell_title_font)

        for cr in self.cr_result:
            r += 1
            r0 = r
            rc = f"A{r}"
            ws[rc] = cr.severity.upper()
            print(f"{rc}, {cr.severity.upper()}")
            xl_format.set_cell(ws[rc])

            rc = f"B{r}"
            lines = 0
            rc_str = f"{cr.msg}"
            lines += 1
            if len(cr.msg) != len(cr.verbose):
                rc_str += '\n'
                rc_str += f"{cr.verbose}"
                lines += len(cr.verbose)/Team().COL_B_WIDTH + 1
                print("lines: " + str(lines))
            ws[rc] = rc_str
            xl_format.set_cell(ws[rc])
            xl_format.set_row(ws, r, 15*lines + 10)
            print(f"{rc}, {rc_str}, {cr.verbose}")

            r += 1
            rc = f"B{r}"
            rc_str = ''
            lines = 0
            for location in cr.locations:
                if lines > 0:
                    rc_str += '\n'
                if len(cr.locations) == 1:
                    rc_str += f"{cr.id}:"
                else:
                    rc_str += f"{location.info}:"
                rc_str += '\n'
                lines += 1
                l = f"{location.filename} line:{location.line}, col:{location.column}"
                rc_str += l
                lines += len(l)/Team().COL_B_WIDTH + 1
            ws[rc] = rc_str
            xl_format.set_cell(ws[rc])
            xl_format.set_row(ws, r, 15*lines + 10)
            print(f"{rc}, {ws[rc]}")

            r += 1
            rc = f"B{r}"
            ws[rc] = f"修复结果："
            xl_format.set_cell(ws[rc])

            mg_rc = f"A{r0}:A{r}"
            ws.merge_cells(mg_rc)

        return r - ori_r

    def cr_result_in_text(self):
        s_text = ''
        s_text += f"{self.name_cn}({self.name_en}):\n"
        for cr in self.cr_result:
            s_text += f"\t{cr.severity.upper()}:\n"
            s_text += f"\t{cr.msg}\n"
            if len(cr.msg) != len(cr.verbose):
                s_text += f"\t{cr.verbose}\n"
            for location in cr.locations:
                if len(cr.locations) == 1:
                    s_text += f"\t\t{cr.id}:\n"
                else:
                    s_text += f"\t\t{location.info}:\n"
                s_text += f"\t\t{location.filename} line:{location.line}, col:{location.column}\n"
            s_text += f"\t修复结果：\n"
            s_text += '\n'
        s_text += '\n\n'

        return s_text

    def kpi_in_text(self):
        report_secondary_seperator = '''------------------------------------------------------------------------'''
        report_summary = '''总结：'''

        kpi_text = self.name_cn + "(" + self.name_en + "):" + '\n'
        kpi_text += self.kpi.fae_bug.get_info() + '\n'
        kpi_text += self.kpi.prot_dev.get_info() + '\n'
        kpi_text += self.kpi.requirement.get_info() + '\n'
        kpi_text += self.kpi.st_bug.get_info() + '\n\n'

        kpi_text += report_summary + '\n'
        kpi_text += "\tWorkload:" + '\n'
        kpi_text += self.kpi.fae_bug.summary + '\n'
        kpi_text += self.kpi.requirement.summary + '\n'
        kpi_text += self.kpi.st_bug.summary + '\n'
        kpi_text += self.kpi.prot_dev.summary + '\n\n'

        kpi_text += report_secondary_seperator + '\n'

        return kpi_text

    def __init__(self):
        self.name_en = ''
        self.name_cn = ''
        self.group = 0  # 'system_development_group: 1', 'application_development_group: 2' and 'tool_development_group: 3'
        self.work_modules = []
        self.work_apps = []
        self.work_gl_apps = []
        self.work_pro_apps = []
        self.cr_result = []
        self.kpi = KPIForOnePerson(self.name_en, self.name_cn)


class Team:
    class TeamGroups:
        def get_group_id_by_name(self, name):
            return self.team_group[name]

        def get_groups_name(self):
            return list(self.team_group.keys())

        def __init__(self):
            self.GROUP_SYSTEM = 1
            self.GROUP_APPLICATION = 2
            self.GROUP_TOOL = 3
            self.team_group = {
                'system_development_group': self.GROUP_SYSTEM,
                'application_development_group': self.GROUP_APPLICATION,
                'tool_development_group': self.GROUP_TOOL
            }

    def save_as_excel(self, select, excel):
        app_r = 1
        sys_r = 1

        if len(self.members) == 0:
            print('No members found.')
            return ''

        wb = Workbook()
        ws_app = wb.active
        ws_app.title = "Application"
        ws_sys = wb.create_sheet("System")

        for mb in self.members:
            if 'cr_result' == select:
                if mb.group == self.tg.GROUP_APPLICATION:
                    app_r += mb.cr_result_in_excel(ws_app, app_r)
                    app_r += 1
                if mb.group == self.tg.GROUP_SYSTEM:
                    sys_r += mb.cr_result_in_excel(ws_sys, sys_r)
                    sys_r += 1

        xl_format.set_column(ws_app, 'A', self.COL_A_WIDTH)
        xl_format.set_column(ws_app, 'B', self.COL_B_WIDTH)
        xl_format.set_column(ws_sys, 'A', self.COL_A_WIDTH)
        xl_format.set_column(ws_sys, 'B', self.COL_B_WIDTH)

        # Save the file
        wb.save(excel)

    def save_as_text(self, select, txt):
        summary_in_text = ''

        if len(self.members) == 0:
            print('No members found.')
            return summary_in_text

        if 'kpi' == select:
            summary_in_text += kpi_head_text()

        for mb in self.members:
            if 'cr_result' == select:
                summary_in_text += mb.cr_result_in_text()
            if 'kpi' == select and mb.group == self.tg.GROUP_APPLICATION:
                summary_in_text += mb.kpi_in_text()

        with open(txt, 'w', encoding='utf-8') as f:
            f.write(summary_in_text)

    def parse_developer(self, node):
        for subNode in node.childNodes:
            if CR_ELEMENT_NODE == subNode.nodeType:
                if 'developer' == subNode.nodeName:
                    m = TeamMember()
                    m.name_en = subNode.getAttribute('name_en')
                    m.name_cn = subNode.getAttribute('name_cn')
                    m.group = self.tg.get_group_id_by_name(node.nodeName)
                    parse_work(m, subNode)
                    m.kpi = KPIForOnePerson(m.name_en, m.name_cn)
                    self.members.append(m)
    
    def member_db_pro(self):
        
        for mb in self.members:
            l = []
            l.append(mb.name_en)
            l.append(mb.name_cn)
            l.append(mb.work_apps)
            l.append(mb.work_gl_apps)
            l.append(mb.work_pro_apps)
            l.append(mb.work_modules)
            l.append(mb.group)
            # print("llll: " + str(l))
            self.member_db_list.append(l)
        # print("222: " + str(self.member_db_list))

    def print_members(self):
        for mb in self.members:
            print(mb.name_cn)
            print(mb.name_en)
            print(mb.group)
            print(mb.work_apps)
            print(mb.work_modules)
            print(mb.work_gl_apps)
            print(mb.work_pro_apps)
        # print(self.member_db_list)

    def init_members(self):
        file = "xml\\team.xml"
        try:
            dom = xml.dom.minidom.parse(file)
        except DOMException:
            raise Exception(file + ' is NOT a well-formed XML file.')

        root = dom.documentElement

        if root.nodeName != 'team':
            raise Exception('XML has no \'Team\' element.')

        # Get attributes of results xx
        results_ver = root.getAttribute('version')
        print(results_ver)

        groups_list = self.tg.get_groups_name()
        for node in root.childNodes:
            if CR_ELEMENT_NODE == node.nodeType:  # 1 is Element
                if node.nodeName in groups_list:
                    self.parse_developer(node)
        
        # self.member_db_pro()
        self.print_members()
    
    def get_level(self, rank):
        '''
        Get the level according to the rank.
        S: 5%
        A: 10%
        B: 50%
        C: 30%
        D: 5%
        '''
        _rank = int(rank)
        total_members = len(self.members)
        if _rank <= int(total_members * 0.05):
            return 'S'
        elif _rank <= int(total_members * 0.15):
            return 'A'
        elif _rank <= int(total_members * 0.65):
            return 'B'
        elif _rank <= int(total_members * 0.95):
            return 'C'
        else:
            return 'D'

    def team_preformance(self, year, season, r_path, fmt, template = "kpi_interview_form_template.docx"):
        '''
        Calculate the team performance.
        '''
        self.members = sorted(self.members, key=lambda mb: mb.kpi.perf.total_score, reverse=True)

        for mb in self.members:
            mb.kpi.perf.pm_score = str(mb.kpi.perf.pm_score)
            mb.kpi.perf.supervisor_score = str(mb.kpi.perf.supervisor_score)
            mb.kpi.perf.total_score = str(mb.kpi.perf.total_score)
            mb.kpi.perf.rank = str(self.members.index(mb) + 1)
            mb.kpi.perf.level = self.get_level(mb.kpi.perf.rank)
            print(mb.name_cn + " " + str(mb.kpi.perf.total_score) + " " + str(mb.kpi.perf.rank) + " " + str(mb.kpi.perf.level))


        docx_template = "kpi_interview_form_template.docx"
        kpi_interview_form_generator(template,  self.tg, year, season, r_path, self.members)

    def __init__(self):
        self.COL_A_WIDTH = 20
        self.COL_B_WIDTH = 120
        self.tg = self.TeamGroups()
        self.members = []
        self.member_db_list = []


software_develop_team = Team()