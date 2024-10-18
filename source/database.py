import os
import time
import openpyxl
import mysql.connector
from mysql.connector import errorcode
from excel_format import ExcelFormat
# from sql import sql_statement as sql_stt

xl_format = ExcelFormat()


class Tables:
    def __init__(self):
        self.name = ''
        self.headline = []


class TeamMembersTable(Tables):
    def get_value(self) -> object:
        print(self.name)
        return sql_stt.get_unique_parameter_list()

    def __init__(self):
        super().__init__()
        self.name = "parameter_table"
        self.headline = ['name_en', 'name_cn', 'work_gv_apps', 'work_gl_apps', 'work_gv_pro_apps','work_module', 'department']
        self.column_values = []


def is_ascii(c):
    if 0 <= ord(c) <= 127:
        return True
    else:
        return False


def is_ascii_str(in_str):
    for c in in_str:
        if not is_ascii(c):
            return False
    return True


class DataBase:
    def __init__(self):
        self.COL_A_WIDTH = 20
        self.COL_B_WIDTH = 120
        self.DATABASE = "demodb"
        self.unit_list = []
        self.db_config = {
            "host": '192.168.10.202',
            "user": 'root',
            "port": '3306',
            "password": 'Sh@nghai9',
            "charset": 'utf8mb4',
            "database": 'demodb'
        }
        self.tables_dict = {
            "team_members": TeamMembersTable()
        }

    def get_table_value(self, table_name, first_line):
        table = self.tables_dict[table_name]
        if first_line:
            return table.headline
        else:
            # return table.get_value(prj)
            return ''

    def mysql_connect(self, attempts=3, delay=2):
        attempt = 1
        while attempt < attempts + 1:
            try:
                return mysql.connector.connect(**self.db_config)
            except (mysql.connector.Error, IOError) as err:
                if attempt == attempts:
                    print("Failed to connect to MySQL")
                    return None
                time.sleep(attempt ** delay)
                attempt += 1

        return None

    def mysql_create_database(self, cursor):
        """
        create database if not exists
        :return: 1 success, 0 fail
        """
        sqlc_creat_database = "{} {} {}".format("CREATE DATABASE", self.DATABASE, "DEFAULT CHARACTER SET \'utf8mb4\'")
        try:
            cursor.execute(sqlc_creat_database)

        except mysql.connector.Error as err:
            print("Failed creating database: {}".format(err))
            return 0
        try:
            sqlc_use_database = "USE {}".format(self.DATABASE)
            cursor.execute(sqlc_use_database)
        except mysql.connector.Error as err:
            print("Database {} does not exists.".format(self.DATABASE))
            if err.errno == errorcode.ER_BAD_DB_ERROR:
                self.mysql_create_database(cursor)
                print("Database {} created successfully.".format(self.DATABASE))
                # cnx.database = self.DATABASE
            else:
                print(err)
                return 0

        return 1

    def mysql_create_table(self, cursor, table):
        """
        If table does not exist, create table
        :param cursor:
        :param table: table name
        :return: 1 success, 0 fail
        """

    def get_unit_table_id(self, in_str):
        print(in_str)
        lower_unit_list = []
        for u in self.unit_list:
            if is_ascii_str(u):
                lower_unit_list.append(u.lower())
            else:
                lower_unit_list.append(u)

        in_str = in_str.lower()
        print("in_str: " + in_str)

        if not is_ascii_str(in_str):
            return self.unit_list.index(in_str) + 1
        elif in_str in lower_unit_list:
            return lower_unit_list.index(in_str.lower()) + 1
        else:
            print("Not found")
            return 0

    def mysql_select(self, cursor, table, columns=None, where=None):

        if columns is None:
            return

        query = "SELECT "
        col = ""
        if len(columns) == 1:
            col += columns[0]
        else:
            for c in columns:
                col += "{}, ".format(c)
            col = col.strip()
            if col.endswith(',') != -1:
                col = col[:-1]
        query += col + " "

        query += "FROM {}".format(table)

        if where is not None:
            print("where")

        print("query: " + query)
        cursor.execute(query, [])
        result = cursor.fetchall()
        t = ()
        for r in result:
            t += r
        self.unit_list = list(t)
        print(table + ": " + str(self.unit_list))

    def mysql_insert(self, cursor, table, rows_l):
        columns = self.get_table_value(table, True)
        values = self.get_table_value(table, False)
        print(columns, values)
        fmt = ''
        cols = '('
        for cs in columns:
            cols += cs
            cols += ','
            fmt += '%s,'
        # discard the last ','
        columns_value = fmt[:-1]
        # discard the last ',' and append a ')'
        columns_name = cols[:-1] + ')'
        print(f"columns_name:{columns_name}  columns_value:{columns_value}")
        print(cols)
        add_columns = ("{} {} ".format("INSERT INTO", table) + columns_name + " VALUES ({})".format(columns_value))
        print(add_columns)
        print("---2---{}, {}".format(len(values), len(columns)))
        if len(values) == len(columns):
            print(tuple(values))
            cursor.execute(add_columns, tuple(values))
        else:
            l = []
            c = 1
            for v in values:
                if isinstance(v, list):
                    l = v
                else:
                    l.append(v)
                print('---3--- ' + str(c) + ': ' + str(l))
                cursor.execute(add_columns, l)
                l.clear()
                c += 1

    def init_prot_info(self):
        conn = self.mysql_connect()
        if conn and conn.is_connected():
            print("Connected to MySQL")
            with conn.cursor() as cursor:
                self.mysql_select(cursor, "team_members", ["name_en,name_cn"])
            conn.close()
        else:
            print("Failed to connect to MySQL")

    def mysql_proc(self, table_name, rows_list):
        print(f"table: {table_name}")
        conn = self.mysql_connect()
        if conn and conn.is_connected():
            print("Connected to MySQL")
            with conn.cursor() as cursor:
                print("get cursor")
                # self.mysql_create_database(cursor)
                # self.mysql_create_table(cursor, table_name)
                self.mysql_insert(cursor, table_name, rows_list)
                conn.commit()
                # cursor.execute(f"SELECT * FROM {table_name}")
                # rows = cursor.fetchall()
                # for row in rows:
                #     print(row)
            conn.close()
        else:
            print("Failed to connect to MySQL")

    def save_as_excel(self, dest_path, file_name, sheet_name):
        excel_full_path = str(os.path.join(dest_path, file_name))

        # wb = Workbook()
        # ws_app = wb.active
        # ws_app.title = sheet_name
        # ws_sys = wb.create_sheet("System")
        try:
            wb = openpyxl.load_workbook(excel_full_path)
        except FileNotFoundError:
            wb = openpyxl.Workbook()

        try:
            ws = wb[sheet_name]
        except KeyError:
            ws = wb.active
            ws.title = sheet_name

        # 获得最大行数
        max_row_num = ws.max_row
        # 获得最大列数
        max_col_num = ws.max_column
        # 将当前行设置为最大行数
        ws._current_row = max_row_num
        print(f"max_row_num: {max_row_num}, max_col_num: {max_col_num}, sheet_name: {sheet_name}")

        # 使用append方法，将行数据按行追加写入
        if max_row_num == 1:
            values = self.get_table_value(sheet_name, True)
            ws.append(values)
        values = self.get_table_value(sheet_name, False)
        ws.append(values)

        # xl_format.set_column(ws_app, 'A', self.COL_A_WIDTH)
        # xl_format.set_column(ws_app, 'B', self.COL_B_WIDTH)

        # Save the file
        wb.save(excel_full_path)


data_base = DataBase()
